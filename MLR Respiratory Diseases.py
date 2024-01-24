#Script for structure the data of PM 2.5
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

#Open the CSV file
file_path = 'PM 2.5_voivodeship.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Process the data
processed_data = {}
for row in range(2, sheet.max_row + 1):  #Row 1 contains headers
    year = sheet.cell(row=row, column=1).value
    voivodeship = sheet.cell(row=row, column=2).value
    avg_time = sheet.cell(row=row, column=7).value
    avg_pm25 = sheet.cell(row=row, column=8).value

    if avg_time == '24g':
        if year not in processed_data:
            processed_data[year] = {'Year': year, 'data': {voivodeship: [avg_pm25]}}
        else:
            if voivodeship not in processed_data[year]['data']:
                processed_data[year]['data'][voivodeship] = [avg_pm25]
            else:
                processed_data[year]['data'][voivodeship].append(avg_pm25)

# Calculate the average of avg_pm25 for each year and voivodeship
averaged_data = {}
for year, data in processed_data.items():
    for voivodeship, pm25_data in data['data'].items():
        avg_pm25 = sum(pm25_data) / len(pm25_data)
        if year not in averaged_data:
            averaged_data[year] = {'Year': year, 'data': {voivodeship: avg_pm25}}
        else:
            if voivodeship not in averaged_data[year]['data']:
                averaged_data[year]['data'][voivodeship] = avg_pm25
            else:
                averaged_data[year]['data'][voivodeship] = (averaged_data[year]['data'][voivodeship] + avg_pm25) / 2

columns = ['Year', 'voivodeship', 'avg_pm25']
avg_list = []

# Iterate through the 'averaged_data' dictionary and append each entry as a new row to the data frame
for year, data in averaged_data.items():
    for voivodeship, avg_pm25 in data['data'].items():
        df_tempt = pd.DataFrame({'Year': [year], 'Voivodeship': [voivodeship], 'Avg_pm25': [avg_pm25]})
        avg_list.append(df_tempt)

avg_df = pd.concat(avg_list, ignore_index=True)

#Script for Linear Regression

# importing modules and packages
import pandas as pd
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
import seaborn as sns

# importing data
doc1 = avg_df
doc2 = pd.read_excel('Diseases of the respiratory system.xlsx', header=0)

merged_data = pd.merge(doc1,doc2, on=['Year', 'Voivodeship'])
print(merged_data)

# Extract the factors X and Y
Y = merged_data['Avg_pm25']
X = merged_data['Respiratory_Diseases']

# Initialize an empty list to store the results
results = []

# Iterate through each unique value in the 'Voivodeship' column
for voivodeship in merged_data['Voivodeship'].unique():
    # Filter the DataFrame to only include rows with the current value in the 'Voivodeship' column
    voivodeship_data = merged_data[merged_data['Voivodeship'] == voivodeship]

    # Calculate the correlation coefficient between X and Y for the filtered DataFrame
    corr = voivodeship_data['Respiratory_Diseases'].corr(voivodeship_data['Avg_pm25'])

    # Fit a regression model using OLS (ordinary least squares) on the filtered DataFrame
    result = smf.ols(formula='Respiratory_Diseases ~ Avg_pm25', data=voivodeship_data).fit()

    # Add the correlation coefficient and regression summary to the results list
    results.append((voivodeship, corr))

# Create a DataFrame from the results list
results_df = pd.DataFrame(results, columns=['Voivodeship','Correlation'])

# Get the maximum value of the y data
max_value = merged_data['Avg_pm25'].max()

# Visualize the relationship between X and Y
plt.figure(figsize=(12, 7))
sns.scatterplot(data=merged_data, y='Avg_pm25', x='Respiratory_Diseases',hue='Voivodeship',palette='deep')
plt.ylabel('Annual average concentration of PM 2.5 (µg/m3)')
plt.xlabel('Death caused by respiratory diseases (Rate)')
plt.title('Scatterplot of Average Concentration of PM 2.5 and Death Caused by Respiratory Diseases from 2013 to 2020', pad=15)
plt.legend(bbox_to_anchor=(1.05, 1), borderaxespad=0., title='Voivodeships')
plt.tight_layout()

# Set the upper limit of the y-axis to the maximum value of the y data
plt.ylim(bottom=0, top=max_value+2)

# Add a horizontal line at y = 5
plt.axhline(y=5, color='black', linestyle='-', linewidth=1)

# Add the word "WHO" at the top of the line
plt.text(0.7, 0.22, 'WHO', transform=plt.gcf().transFigure, fontsize=12, fontweight='bold', color='black', verticalalignment='bottom')

# Add light grey color for the area above the line on y-axis value = 5
plt.axhspan(5, max_value+2, color='lightgrey', alpha=0.3)

plt.show()

# Print the DataFrame
print(results_df)

# Save the DataFrame to a CSV file
results_df.to_excel('output_respiratory.xlsx', index=False)
